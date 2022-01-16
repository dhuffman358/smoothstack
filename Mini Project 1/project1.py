from ctypes.wintypes import tagMSG
from datetime import date
from matplotlib.pyplot import contour
import openpyxl
import logging
import pathlib
import datetime


#Program to import data from a specific Expedia monthly report and give information from the corresponding month and year.

#Set up logger
logging.basicConfig(filename="report_log",
                    format='%(asctime)s %(message)s',
                    filemode='w')
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
#Set up dictionaries to convert between month name, three letter month abbreviation, and number of month
months = {'JAN':'january', 'FEB':"february", 'MAR':'march', 'APR': 'april', 'MAY':'may', 'JUN' : 'june',
        'JUL': 'july,','AUG':'august','SEP':'september','OCT':'october','NOV':'november','DEC':'december'}
month = ''
month_number = {'JAN':1, 'FEB':2, 'MAR':3, 'APR': 4, 'MAY':5, 'JUN' : 6,
                    'JUL': 7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
#Set up years to limit input
years = range(1990, 2100)
year = 0
exit = False
while not exit: #loop to return to top to get input again if file does not exist
    #loop to get month and year desired of report
    while True:
        month = input('Input month wanted as three letters:   (Type QUIT to quit)\n').upper()
        if month == 'QUIT':
            exit = True
            logging.info('Program ended by user.')
            break
        elif month not in months:
            logging.info('Wrong month data inputted, trying again')
            print("Please try again using first three letters of month.")
            continue
        else:
            logging.info('Month is ' + months[month])
        break
    if month == 'QUIT':
        break
    while True:  
        try:
            year = int(input("Input year wanted in format xxxx:\n"))
        except ValueError:
            print("Please try again using four digit year between 1990 and 2099.")
            logging.error("Non-integer input for year, trying again")
            continue
        if year not in years:
            logging.info('Wrong year data inputted, trying again')
            print("Please try again using four digit year between 1990 and 2099.")
            continue
        logging.info('Year is ' + str(year))
        break
    local_folder = pathlib.Path(__file__).parent.resolve().as_uri()
    report_filename =  local_folder[8:] + '/expedia_report_monthly_' + months[month] +'_' + str(year) +'.xlsx'
    logging.info('File name is ' + report_filename)
    try:
        expedia_workbook = openpyxl.load_workbook(report_filename)
    except FileNotFoundError:
        input('File not found for that month and year. Ensure script and files are in same folder. Returning to data input.')
        logging.error("File inputted not found.")
        continue
    sheet1 = expedia_workbook['Summary Rolling MoM']
    sheet2 = expedia_workbook['VOC Rolling MoM']
    date_code = datetime.datetime(year, month_number[month], 1, 0, 0)
    target_row = ()
    target_col = ()
    for row in sheet1.iter_rows(min_row = 1, max_row =15, values_only=True):
        try:
            if row[0].year == date_code.year and row[0].month == date_code.month:
                target_row = row
                logger.info(f" Data for selected year on sheet \'Summary Rolling MoM\' is {target_row[0:5]}")
        except:
            continue
    for col in sheet2.iter_cols(min_row =1, min_col=1, values_only=True):
            if type(col[0]) == type(''):
                if col[0].upper().find(month) != -1 or col[0].upper()[0:2] == month: #checks in case data is a word
                    target_col = col
                    break #data is arranged with leftmost as most recent, if title of column is a word just want the first march
            if type(col[0]) == type(date_code):
                if col[0].month == date_code.month: #checks in case data is in datetime format
                    target_col = col
                    break
    promoters = ''
    passives = ''
    detractors = ''
    if target_col[3] > 200:
        promoters = "good"
    else:
        promoters = 'bad'
    if target_col[5] > 100:
        passives = 'bad'
    else:
        passives = 'good'
    if target_col[7] > 100:
        detractors = 'bad'
    else:
        detractors = 'good'
    column_data = (promoters, passives, detractors, target_col[15], target_col[15], target_col[18])
    output_sheet1 = f"Calls Offered: {target_row[1]} \nAbandon after 30s: {target_row[2] *100}% \nFCR: {target_row[3]*100}% \nDSAT: {target_row[4]*100}% \nCSAT: {target_row[5]*100}%"
    output_sheet2 = f"Promoters: {column_data[0]}\nPassives: {column_data[1]}\nDetractors: {column_data[2]}\nOverall NPS %: {column_data[3]*100}%\nSat with Agent %: {column_data[4]*100}%\nDSAT with Agent % {column_data[5]*100}%"
    logging.info(output_sheet1)
    logging.info(output_sheet2)
    print(f"{output_sheet1}\n{output_sheet2}")
    another_report = input("Type REPEAT to get the data from another report, otherwise enter to quit:\n")
    if  another_report.upper() != 'REPEAT':
        break