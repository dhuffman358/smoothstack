from multiprocessing.sharedctypes import Value
import openpyxl
import logging
import pathlib
import datetime


"""Function to import data from a specific Expedia monthly report and give information from the corresponding month and year. Takes a filename as input.
File must be in format expedia_report_monthly_xxxx_yyyy.xlsx where xxxx is the full name month (eg, january) and yyyy is four digit year (eg, 2018)
Returns string of information from report if input is valid. 
Raises ValueError if filename is in wrong format and FileNotFound if the file does not exist."""
def get_report_information(report_filename):
    #Set up logger
    logging.basicConfig(filename="mini_project/get_report_information_log", #todo add current date to logfile name
                        format='%(asctime)s %(message)s',
                        filemode='w')
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logging.info(f"Running script {pathlib.Path(__file__).resolve().as_uri()}")

    #dictionary to convert filename month into number for datetime
    month_number = {'january':1, 'february':2, 'mar':3, 'april': 4, 'may':5, 'june' : 6,
                    'july': 7,'august':8,'september':9,'october':10,'november':11,'december':12}

    #logic to ensure that the file name is in the correct format: expedia_report_monthly_january_2018.xlsx and extract month and year
    try:
        month_and_year = pathlib.Path(report_filename).stem
        print(month_and_year)
        month = month_and_year[23:-5].lower()
        year = int(month_and_year[-4:])
    except IndexError:
        logging.error('Value Error: Improper file format inputted into get_report_information: filename too short')
        raise ValueError
    print(month)
    print(year)
    if month not in month_number or year not in range(1996,2100):
        logging.error('Value Error: Improper file format inputted into get_report_information: month or year not found')
        raise ValueError
    if report_filename[-4:] != 'xlsx':
        logging.error('Value error: Not an xlsx file.')
        raise ValueError
    logging.info(f"Month is {month} and year is {year}")
    print(report_filename)
    #Loading file and getting the two relevant sheets from the spreadsheet
    try:
        expedia_workbook = openpyxl.load_workbook(report_filename)
    except FileNotFoundError:
        logging.error('File not found for that month and year. Ensure script and files are in same folder.')
        raise FileNotFoundError
    logging.info(f"File loaded as {report_filename}")
    sheet1 = expedia_workbook['Summary Rolling MoM']
    sheet2 = expedia_workbook['VOC Rolling MoM']
    date_code = datetime.datetime(year, month_number[month], 1, 0, 0) #requested report year and month in datetime format

    #Finding the target data
    target_row = ()
    target_col = ()
    for row in sheet1.iter_rows(min_row = 1, max_row =15, values_only=True):
        try:
            if row[0].year == date_code.year and row[0].month == date_code.month:
                target_row = row
        except:
            continue
    for col in sheet2.iter_cols(min_row =1, min_col=1, values_only=True):
            if type(col[0]) == type(''):
                if col[0].upper().find(month) != -1 or col[0].upper()[0:2] == month: #checks in case data is a word
                    target_col = col
                    break #data is arranged with leftmost as most recent, if title of column is a word just want the first march
            if type(col[0]) == type(date_code):
                if col[0].month == date_code.month: #checks in case data is in datetime format
                    target_col = col
                    break
    #Assembling the row and column data for returns in case needs integration into bigger system
    try:
        row_data    = (target_row[1],  #calls offered
                    target_row[2]*100, #abandon after
                    target_row[3]*100, #FCR
                    target_row[4]*100, #DSAT
                    target_row[5]*100  #CSAT
                    )
        column_data = ((lambda x : 'good' if x >= 200 else 'bad')(target_col[3]), #promoters evaluation
                    (lambda x : 'good' if x <= 100 else 'bad')(target_col[5]), #passives
                    (lambda x : 'good' if x <= 100 else 'bad')(target_col[7]), #detractors
                    target_col[15]*100, #Overall NPS
                    target_col[15]*100, #Sat with Agent
                    target_col[18]*100  #Detractors
                    )
    except:
        logging.error("get_report_information: File data in wrong format/corrupted.")
        raise ValueError

    logging.info(f"Data from sheet \'Summary Rolling MoM\': {row_data}")
    logging.info(f"Data from sheet \'VOC Rolling Mom\': {column_data}")
    output_sheet1 = f"Calls Offered: {row_data[0]} \nAbandon after 30s: {row_data[1]}% \nFCR: {row_data[2]}% \nDSAT: {row_data[3]}% \nCSAT: {row_data[4]}%\n"
    output_sheet2 = f"Promoters: {column_data[0]}\nPassives: {column_data[1]}\nDetractors: {column_data[2]}\nOverall NPS %: {column_data[3]}%\nSat with Agent %: {column_data[4]}%\nDSAT with Agent % {column_data[5]}%"
    return output_sheet1 + output_sheet2
    
def console_input():
    #Set up dictionaries to convert between month name, three letter month abbreviation, and number of month (for datetimes)
    months = {'JAN':'january', 'FEB':"february", 'MAR':'march', 'APR': 'april', 'MAY':'may', 'JUN' : 'june',
            'JUL': 'july,','AUG':'august','SEP':'september','OCT':'october','NOV':'november','DEC':'december'}
    month = ''

    #Set up years to limit input
    years = range(1996, 2100)
    year = 0
    exit = False
    while not exit: #loop to return to top to get input again if file does not exist or to repeat for another file   
    #loops to get month and year desired of report
        while True:
            month = input('Input month wanted as three letters:   (Type QUIT to quit)\n').upper()
            if month == 'QUIT':
                exit = True
                break
            elif month not in months:
                print("Please try again using first three letters of month.")
                continue
            break
        if month == 'QUIT':
            break
        while True:  
            try:
                year = int(input("Input year wanted in format xxxx:\n"))
            except ValueError:
                print("Please try again using four digit year between 1996 and 2099.")
                continue
            if year not in years:
                print("Please try again using four digit year between 1996 and 2099.")
                continue
            break
        #Generating filename as per user input
        local_folder = pathlib.Path(__file__).parent.resolve().as_uri() 
        report_filename =  local_folder[8:] + '/expedia_report_monthly_' + months[month] +'_' + str(year) +'.xlsx'
        try:
            print(get_report_information(report_filename))
        except ValueError:
            print("Filename error or data in file error. Try again?")
        except FileNotFoundError:
            print('File not found. Try again?')
        another_report = input("Type REPEAT to get the data from another report, otherwise enter to quit:\n")
        if  another_report.upper() != 'REPEAT':
            logging.info("Program complete and ended.")
            break
if __name__ == '__main__':
    console_input()